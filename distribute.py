from openpyxl import load_workbook
import pandas as pd
from dotenv import load_dotenv
import os
from datetime import datetime

load_dotenv()

REGISTRATIONS_FOLDER_PATH = os.environ.get("REGISTRATIONS_FOLDER_PATH")

# NOTE: Make sure to keep this updated
EXCELS = {
    "CBBD": ("Circular Bioeconomy Business Development - Registrations.xlsx"),
    "CACE": ("Climate Action and Community Engagement - Registrations.xlsx"),
    "CVA": ("Climate Vulnerability and Adaptation - Registrations.xlsx"),
    "CNR": ("Co-Management of Natural Resources - Registrations.xlsx"),
    "CSRP": ("Communication Strategies for Resource Practitioners - Registrations.xlsx"),
    "EFO": ("Environmental Footprints of Organizations - Registrations.xlsx"),
    "FSTB": ("Fire Safety for Timber Buildings - Registrations.xlsx"),
    "FCM": ("Forest Carbon Management - Registrations.xlsx"),
    "FHM": ("Forest Health Management - Registrations.xlsx"),
    "HTC": ("Hybrid Timber Construction - Registrations.xlsx"),
    "SMS": ("Strategic Management for Sustainability - Registrations.xlsx"),
    "TWS": ("Tall Wood Structures - Registrations.xlsx"),
    "ZCBS": ("Zero Carbon Building Solutions - Registrations.xlsx"),
}

# This is where to expect the header to be in the excel, necessary for finding the right column for data
HEADER_ROW = 2

def add_date_to_filename(filename):
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    # Split the filename into name and extension
    name, extension = os.path.splitext(filename)

    # Return the filename with date appended
    return f"{name}_{date_str}{extension}"

"""
Steps to distribute data for each enrollment entry
1. Iterate through new enrollment data (must be a pandas dataframe)
Enrollment data must contain the following columns:
student_name_0 -> Full name, student_name_1 -> Email address, product_name_0 (contains program code and session)
Then using their email address retrieve the following data from user_data (take the last match, potentially duplicate emails if data changed):
custom_fields_organization -> Organization
custom_fields_title -> Title
custom_fields_phone-number -> Phone Number
custom_fields_mailing-address -> Mailing Address (Only check for CNR)
custom_fields_indigenous-self-declaration -> Self-Identify as Indigenous? (Only check for CNR)

Also pull the following data from processed_data by searching for student's email in processed_data
Grant amount to give -> Received FSG?, Grant Amount

2. Use account_name + product_name_0 to find the correct sheet to use (excel name and sheet session)
3. Check sheet if user (based on email) already exists save the row index if yes 
4. Append the data to the end of the sheet, or write to the exiting row
"""

def extract_user_data(row, user_data_row, user_grant_row):
    """ This puts combines the data from the various sheets into the format we want """
    email = row['student_name_1'].split(' ')[2]

    data = {
        'Full Name': row['student_name_0'],
        'Email Address': email,
    }

    if not user_data_row.empty:
        extra = {
            'Organization': user_data_row['custom_fields_organization'].values[0],
            'Title': user_data_row['custom_fields_title'].values[0],
            'Phone Number': user_data_row['custom_fields_phone-number'].values[0],
            'Mailing Address': user_data_row['custom_fields_mailing-address'].values[0],
            'Self-Identify as Indigenous?': 'Yes' if user_data_row['custom_fields_indigenous-self-declaration'].values[0].lower().strip() == '1' else 'No',
        }
        data.update(extra)

    if not user_grant_row.empty:
        extra = {
            'Received FSG?': 'Yes',
            'Grant Amount Received': user_grant_row['Grant amount to give'].values[0]
        }
        data.update(extra)

    return data

def find_sheet(row):
    """ This finds the correct sheet based on the user's program, assumes the correct sheet and excel exists """
    course_code = row['account_name'].split(" ")[0]
    course_info = row['product_name_0'].split(" ")
    course_session = " ".join(course_info[-2:])

    excel_path = REGISTRATIONS_FOLDER_PATH + EXCELS[course_code]
    workbook = load_workbook(filename=excel_path)
    return (workbook[course_session], excel_path, workbook)

def search_email_in_sheet(sheet, email):
    """ Return the row index where email is found, -1 if not found """
    email_column = None

    for cell in sheet[HEADER_ROW]:
        if cell.value == 'Email Address':
            email_column = cell.column_letter
            break

    if email_column is None:
        return -1

    for row_idx, cell in enumerate(sheet[email_column], start=1):
        if cell.value and cell.value.lower().strip() == email:
            return row_idx

    return -1

def find_empty_row(sheet):
    """ Starting from row 3 of the sheet, find a row where the first column is empty"""
    for row_index, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        if row[0].value is None:
            return row_index
    return sheet.max_row + 1

def insert_or_append_row(sheet, data, existing_row):
    """ If the row exists (not -1) then add data to columns that are empty, else append to end of sheet """
    row = existing_row if existing_row != -1 else None
    if row is None:
        row = find_empty_row(sheet)
        
    for col_header, value in data.items():
        col_index = None
        for cell in sheet[HEADER_ROW]: # Find the column where the current header is
            if cell.value == col_header:
                col_index = cell.column
                break

        if col_index is not None:
            target_cell = sheet.cell(row=row, column=col_index)
            if target_cell.value is None:
                target_cell.value = value
            
def distribute_enrollment_data(df_enrollment, path_to_user_data, path_to_grant_data):
    """ Loops through all the enrollment users, and distributes their data to the correct sheet """
    df_user_data = pd.read_excel(path_to_user_data)
    df_grant_data = pd.read_excel(path_to_grant_data)
    all_rows = []

    for _, row in df_enrollment.iterrows():
        # Search for the row in user_data based on student_name_1 and email inside student_name_1
        user_data_row = df_user_data[df_user_data['student_name_1'].str.lower().str.strip() == row['student_name_1'].lower().strip()].tail(1)
        user_email = row['student_name_1'].split(' ')[2].lower().strip() 
        user_grant_row = df_grant_data[df_grant_data['Email'].str.lower().str.strip() == user_email].tail(1)
        data = extract_user_data(row, user_data_row, user_grant_row)
        # 2: find the correct sheet to use
        (sheet, excel_path, workbook) = find_sheet(row)

        # 3: Check if email already in sheet
        existing_row = search_email_in_sheet(sheet, user_email)

        # 4: Insert data at the end or write to the existing row
        insert_or_append_row(sheet, data, existing_row)
        data["Excel Path"] = excel_path.split("/")[-1]
        all_rows.append(data)
        workbook.save(excel_path)
    
    df = pd.DataFrame(all_rows)
    save_path = add_date_to_filename(os.environ.get('ENROLLMENTS_HISTORY_PATH'))
    df.to_excel(save_path, index=False)
    print("SAVED ENROLLMENTS THAT WILL BE DISTRIBUTED TO", save_path)    

    print("DONE DISTRIBUTING DATA")

if __name__ == '__main__':
    # This is mainly for testing, call python get_data.py instead
    df = pd.read_excel(os.environ.get("RAW_DATA_PATH_ENROLLMENTS"))
    distribute_enrollment_data(df, os.environ.get("RAW_DATA_PATH_USERS"), os.environ.get("PROCESSED_DATA_PATH"))